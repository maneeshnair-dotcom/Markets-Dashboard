# -*- coding: utf-8 -*-
"""
zerodha_analyzer.py
─────────────────────────────────────────────────────────────────────────────
Zerodha (Kite Connect) data-source variant of nifty50_analyzer.py.

Same indicator engine, same output columns, same Excel export as
nifty50_analyzer.py — the only thing that changes is where the OHLCV data
comes from:

    nifty50_analyzer.py : yf.download(tickers, period=..., interval=...)
    zerodha_analyzer.py  : kite.historical_data(instrument_token, ...)

This file does NOT handle Kite login/authentication itself (that's a
Streamlit-UI concern — see dashboard_zerodha.py, adapted from the reference
app(1).py you provided). Every function here takes an already-authenticated
`kite` client (a kiteconnect.KiteConnect instance with set_access_token()
already called) as a parameter.

ONE TICKER LIST ONLY
──────────────────────
Per your request, there's no multi-list registry here (unlike
nifty50_analyzer.py's Nifty/INT/BSE/BANKNIFTY) — just the single `FnO` list
below. Edit it directly to add/remove instruments.

INSTRUMENT RESOLUTION — IMPORTANT CAVEAT
──────────────────────────────────────────
Kite doesn't have a plain "spot price" instrument for everything in your
list — only equities trade as a simple spot instrument. This module
resolves each name in FnO to an actual fetchable instrument as follows:

  • Index names (NIFTY, BANKNIFTY, SENSEX, ...) → the NSE/BSE "INDICES"
    segment instrument, via the INDEX_ALIASES map below (Kite's own
    tradingsymbols for indices don't match the short F&O names — e.g.
    Kite lists "NIFTY 50", not "NIFTY"). I've filled in the aliases I'm
    confident about; double-check them against your own
    kite.instruments("NSE") / kite.instruments("BSE") dump before relying
    on this in production, since I can't verify these against a live Kite
    session from here. "FOCIT" in particular has no standard public index
    I could confidently map — it's left unresolved (skipped with a
    warning) until you add the correct tradingsymbol to INDEX_ALIASES.

  • Currency pairs (EURINR, GBPINR, JPYINR, USDINR) and commodities
    (COPPER, CRUDEOIL, GOLD, NATGASMINI, NATURALGAS, SILVER, SILVERM,
    ZINC) don't trade as a spot instrument on Kite at all — they're
    derivatives-only. This module uses the nearest-expiry (front-month)
    FUTURES contract's own OHLCV series as a continuous proxy. Note this
    means the price series jumps at each monthly rollover (no continuous-
    contract stitching is done) — acceptable for signal generation, but
    keep it in mind if you chart raw price levels across a rollover.

  • Everything else (equity F&O stocks) → the plain NSE equity spot
    instrument, matched by tradingsymbol.

RATE LIMITS
────────────
Kite's historical-data endpoint is rate-limited (~3 requests/second on
standard plans). fetch_fno_data() sleeps between each per-instrument call
and retries once on transient failures, same approach as your reference
app(1).py. Fetching the full ~190-symbol FnO list will take a few minutes
end to end — this is inherent to Kite's per-symbol historical API, there's
no batched multi-symbol historical endpoint like yfinance's.
"""
import time
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from numpy.lib.stride_tricks import sliding_window_view

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print("⚠  openpyxl not found — Excel export disabled.  Run:  pip install openpyxl")

IST = ZoneInfo("Asia/Kolkata")


def now_ist() -> datetime:
    """Naive IST datetime — Kite's API always interprets datetimes as IST,
    but hosted servers often run in UTC, so every 'now' goes through this."""
    return datetime.now(IST).replace(tzinfo=None)


# ── The ONE ticker list (edit this directly to add/remove instruments) ────────
FnO = [
    "NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "NIFTYNXT50",
    "BANKEX", "FOCIT", "SENSEX", "SENSEX50",
    "EURINR", "GBPINR", "JPYINR", "USDINR",
    "COPPER", "CRUDEOIL", "GOLD",
    "NATGASMINI", "NATURALGAS", "SILVER", "SILVERM", "ZINC",
    "360ONE", "ABB", "ABCAPITAL", "ADANIENSOL", "ADANIENT", "ADANIGREEN",
    "ADANIPORTS", "ADANIPOWER", "ALKEM", "AMBER", "AMBUJACEM", "ANGELONE",
    "APLAPOLLO", "APOLLOHOSP", "ASHOKLEY", "ASIANPAINT", "ASTRAL", "AUBANK",
    "AUROPHARMA", "AXISBANK", "BAJAJ-AUTO", "BAJAJFINSV", "BAJAJHLDNG",
    "BAJFINANCE", "BANDHANBNK", "BANKBARODA", "BANKINDIA", "BDL", "BEL",
    "BHARATFORG", "BHARTIARTL", "BHEL", "BIOCON", "BLUESTARCO", "BOSCHLTD",
    "BPCL", "BRITANNIA", "BSE", "CAMS", "CANBK", "CDSL", "CGPOWER",
    "CHOLAFIN", "CIPLA", "COALINDIA", "COCHINSHIP", "COFORGE", "COLPAL",
    "CONCOR", "CROMPTON", "CUMMINSIND", "DABUR", "DALBHARAT", "DELHIVERY",
    "DIVISLAB", "DIXON", "DLF", "DMART", "DRREDDY", "EICHERMOT", "ETERNAL",
    "FEDERALBNK", "FORCEMOT", "FORTIS", "GAIL", "GLENMARK", "GMRAIRPORT",
    "GODFRYPHLP", "GODREJCP", "GODREJPROP", "GRASIM", "GVT&D", "HAL",
    "HAVELLS", "HCLTECH", "HDFCAMC", "HDFCBANK", "HDFCLIFE", "HEROMOTOCO",
    "HINDALCO", "HINDPETRO", "HINDUNILVR", "HINDZINC", "HYUNDAI",
    "ICICIBANK", "ICICIGI", "ICICIPRULI", "IDEA", "IDFCFIRSTB", "IEX",
    "INDHOTEL", "INDIANB", "INDIGO", "INDUSINDBK", "INDUSTOWER", "INFY",
    "INOXWIND", "IOC", "IREDA", "IRFC", "ITC", "JINDALSTEL", "JIOFIN",
    "JSWENERGY", "JSWSTEEL", "JUBLFOOD", "KALYANKJIL", "KAYNES", "KEI",
    "KFINTECH", "KOTAKBANK", "KPITTECH", "LAURUSLABS", "LICHSGFIN", "LICI",
    "LODHA", "LT", "LTF", "LTM", "LUPIN", "M&M", "MANAPPURAM", "MANKIND",
    "MARICO", "MARUTI", "MAXHEALTH", "MAZDOCK", "MCX", "MFSL", "MOTHERSON",
    "MOTILALOFS", "MPHASIS", "MUTHOOTFIN", "NAM-INDIA", "NATIONALUM",
    "NAUKRI", "NBCC", "NESTLEIND", "NHPC", "NMDC", "NTPC", "NYKAA",
    "OBEROIRLTY", "OFSS", "OIL", "ONGC", "PAGEIND", "PATANJALI", "PAYTM",
    "PERSISTENT", "PETRONET", "PFC", "PGEL", "PHOENIXLTD", "PIDILITIND",
    "PIIND", "PNB", "PNBHOUSING", "POLICYBZR", "POLYCAB", "POWERGRID",
    "POWERINDIA", "PREMIERENE", "PRESTIGE", "RADICO", "RBLBANK", "RECLTD",
    "RELIANCE", "RVNL", "SAIL", "SBICARD", "SBILIFE", "SBIN", "SHREECEM",
    "SHRIRAMFIN", "SIEMENS", "SOLARINDS", "SONACOMS", "SRF", "SUNPHARMA",
    "SUPREMEIND", "SUZLON", "SWIGGY", "TATACONSUM", "TATAELXSI",
    "TATAPOWER", "TATASTEEL", "TCS", "TECHM", "TIINDIA", "TITAN", "TMPV",
    "TORNTPHARM", "TRENT", "TVSMOTOR", "ULTRACEMCO", "UNIONBANK",
    "UNITDSPR", "UNOMINDA", "UPL", "VBL", "VEDL", "VMM", "VOLTAS",
    "WAAREEENER", "WIPRO", "YESBANK", "ZYDUSLIFE",
]

# ── Index name -> Kite's actual INDICES-segment tradingsymbol ─────────────────
# CHECK THESE against your own kite.instruments("NSE")/kite.instruments("BSE")
# before relying on them — I could not verify against a live Kite session.
INDEX_ALIASES = {
    "NIFTY":      "NIFTY 50",
    "BANKNIFTY":  "NIFTY BANK",
    "FINNIFTY":   "NIFTY FIN SERVICE",
    "MIDCPNIFTY": "NIFTY MIDCAP SELECT",
    "NIFTYNXT50": "NIFTY NEXT 50",
    "SENSEX":     "SENSEX",
    "BANKEX":     "BANKEX",
    "SENSEX50":   "BSE SENSEX 50",
    # "FOCIT": "..."   <- add the correct Kite tradingsymbol here if you have it
}

# ── Currency pairs & commodities: no spot instrument, use front-month future ──
FUTURES_UNDERLYINGS = {
    "EURINR", "GBPINR", "JPYINR", "USDINR",
    "COPPER", "CRUDEOIL", "GOLD",
    "NATGASMINI", "NATURALGAS", "SILVER", "SILVERM", "ZINC",
}

# Exchanges/segments to pull the instrument master from.
INSTRUMENT_SEGMENTS = ["NSE", "BSE", "CDS", "MCX"]

# ── Dashboard interval label -> Kite historical_data interval ─────────────────
# Kite has no native 4h or 1wk bar — built by resampling 60minute/day.
KITE_INTERVAL_MAP = {
    "15m":  ("15minute", None),
    "30m":  ("30minute", None),
    "1h":   ("60minute", None),
    "4h":   ("60minute", "4h"),
    "1d":   ("day", None),
    "1wk":  ("day", "1W"),
}
# Kite's historical API also enforces its own max lookback per interval
# (roughly: minute-family intervals cap out around 60-100 days per request,
# day interval allows years). If you hit a "date range too long" error from
# Kite for a given n_days/interval combination, verify the current limit in
# Kite Connect's docs and either shorten n_days or add request-chunking here.


# ── Instrument resolution ──────────────────────────────────────────────────────
def get_instrument_master(kite) -> pd.DataFrame:
    """Fetch and combine the instrument list for every segment in
    INSTRUMENT_SEGMENTS. Cache this yourself (e.g. via st.cache_data with a
    ~1hr TTL in the dashboard) — this call is relatively heavy."""
    frames = []
    for seg in INSTRUMENT_SEGMENTS:
        try:
            d = pd.DataFrame(kite.instruments(seg))
            if not d.empty:
                frames.append(d)
        except Exception as e:
            print(f"  ⚠ Could not fetch {seg} instrument list: {e}")
    if not frames:
        return pd.DataFrame()
    combined = pd.concat(frames, ignore_index=True)
    combined["expiry"] = pd.to_datetime(combined["expiry"], errors="coerce")
    return combined


def resolve_instrument(instruments_df: pd.DataFrame, name: str):
    """
    Resolve one FnO list entry to (instrument_token, tradingsymbol, exchange).
    Returns (None, None, None) if it can't be found — caller should skip
    and warn rather than crash the whole run over one bad symbol.
    """
    if instruments_df.empty:
        return None, None, None

    if name in INDEX_ALIASES:
        alias = INDEX_ALIASES[name]
        match = instruments_df[
            (instruments_df["segment"] == "INDICES") &
            (instruments_df["tradingsymbol"] == alias)
        ]
        if match.empty:
            return None, None, None
        row = match.iloc[0]
        return row["instrument_token"], row["tradingsymbol"], row["exchange"]

    if name in FUTURES_UNDERLYINGS:
        today = pd.Timestamp(now_ist().date())
        match = instruments_df[
            (instruments_df["name"] == name) &
            (instruments_df["instrument_type"] == "FUT") &
            (instruments_df["expiry"] >= today)
        ].sort_values("expiry")
        if match.empty:
            return None, None, None
        row = match.iloc[0]  # nearest (front-month) expiry
        return row["instrument_token"], row["tradingsymbol"], row["exchange"]

    # Plain equity — NSE spot
    match = instruments_df[
        (instruments_df["exchange"] == "NSE") &
        (instruments_df["segment"] == "NSE") &
        (instruments_df["tradingsymbol"] == name)
    ]
    if match.empty:
        return None, None, None
    row = match.iloc[0]
    return row["instrument_token"], row["tradingsymbol"], row["exchange"]


def _resample_candles(df: pd.DataFrame, rule: str) -> pd.DataFrame:
    """Roll up finer candles into coarser ones (60minute->4h, day->1W).
    Bins anchor to local midnight, so an NSE 09:15-15:30 session lines up
    into two real 4h candles: 08:00-12:00 and 12:00-16:00."""
    if df.empty:
        return df
    indexed = df.set_index(pd.to_datetime(df["Date"]).dt.tz_localize(None))
    agg = {"OPEN": "first", "HIGH": "max", "LOW": "min", "CLOSE": "last", "VOLUME": "sum"}
    out = indexed.resample(rule).agg(agg)
    out = out.dropna(subset=["OPEN", "HIGH", "LOW", "CLOSE"], how="all")
    return out.reset_index().rename(columns={"index": "Date"})


def _historical_with_retry(kite, token, from_date, to_date, interval, retries=1):
    for attempt in range(retries + 1):
        try:
            return kite.historical_data(token, from_date, to_date, interval)
        except Exception:
            if attempt < retries:
                time.sleep(1.0)
                continue
            raise


# ── Indicator Functions (identical to nifty50_analyzer.py) ────────────────────

def calculate_rsi(series, window=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0).rolling(window).mean()
    loss = -delta.where(delta < 0, 0).rolling(window).mean()
    rs = gain / loss
    return 100 - (100 / (1 + rs))


def calculate_stochastic(df, k_window=12, d_window=5):
    low_min  = df['LOW'].rolling(k_window).min()
    high_max = df['HIGH'].rolling(k_window).max()
    k = 100 * ((df['CLOSE'] - low_min) / (high_max - low_min))
    d = k.rolling(d_window).mean()
    return k.rename('Stoch_K'), d.rename('Stoch_D')


def calculate_bollinger_bands(series, window=20, std_dev=2):
    middle = series.rolling(window).mean()
    std    = series.rolling(window).std()
    upper  = middle + std_dev * std
    lower  = middle - std_dev * std
    width  = ((upper - lower) / middle) * 100
    return middle, upper, lower, width


def bollinger_position(close, upper, lower):
    pos = pd.Series("Inside", index=close.index, dtype=str)
    pos[close > upper] = "Upper_Breakout"
    pos[close < lower] = "Lower_Breakout"
    return pos


def volume_analysis(df):
    """Note: index instruments (NIFTY 50, SENSEX, etc.) report zero/NaN
    volume on Kite — Volume_Trend/Volume_Signal will just be uninformative
    for those rows, everything else still works."""
    df['Volume_SMA']   = df['VOLUME'].rolling(20).mean()
    df['Volume_Ratio'] = (df['VOLUME'] / df['Volume_SMA']).round(2)
    df['OBV']          = (np.sign(df['CLOSE'].diff()) * df['VOLUME']).fillna(0).cumsum()
    df['Volume_Trend'] = np.where(df['VOLUME'] > df['Volume_SMA'], "Rising", "Falling")
    df['Volume_Signal'] = ""
    strong_vol = (df['Volume_Ratio'] > 1.8)
    df.loc[strong_vol & (df['CLOSE'] > df['OPEN']), 'Volume_Signal'] = "Strong_Buy_Vol"
    df.loc[strong_vol & (df['CLOSE'] < df['OPEN']), 'Volume_Signal'] = "Strong_Sell_Vol"
    return df


def simple_gann_time_signals(df):
    bar = np.arange(1, len(df) + 1)
    is_cycle = (bar % 21 == 0) | (bar % 90 == 0)
    df['Gann_Time'] = np.where(is_cycle, "Time_Cycle", "")
    return df


def calculate_gann_targets(df, zone_tolerance=0.018):
    close  = df['CLOSE']
    sqrt_p = np.sqrt(close.clip(lower=0.01))

    upside   = np.ceil(sqrt_p)  ** 2
    downside = np.floor(sqrt_p) ** 2

    df['Gann_Resistance']  = upside.round(2)
    df['Gann_Support'] = downside.round(2)

    up_ratio   = (close - upside).abs()   / upside
    down_ratio = (close - downside).abs() / downside

    df['Gann_Reversal_Zone'] = np.select(
        [up_ratio < zone_tolerance, down_ratio < zone_tolerance],
        ['Resistance', 'Support'],
        default=''
    )

    if len(df) > 0:
        df.iloc[0, df.columns.get_loc('Gann_Resistance')]       = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Support')]      = np.nan
        df.iloc[0, df.columns.get_loc('Gann_Reversal_Zone')]= ""
    return df


def detect_stochastic_divergence(df, window=20):
    price = df['CLOSE']
    stoch = df.get('Stoch_D', pd.Series(np.nan, index=df.index))
    span  = window + 1

    roll_min_p = price.rolling(span, min_periods=1).min()
    roll_max_p = price.rolling(span, min_periods=1).max()
    roll_min_s = stoch.rolling(span, min_periods=1).min()
    roll_max_s = stoch.rolling(span, min_periods=1).max()

    bullish = (price <= roll_min_p * 1.001) & (stoch > roll_min_s * 1.02)
    bearish = (price >= roll_max_p * 0.999) & (stoch < roll_max_s * 0.98)

    div   = pd.Series("", index=df.index, dtype=str)
    valid = np.arange(len(df)) >= window
    div[bullish & valid] = "Bullish_Div"
    div[bearish & valid] = "Bearish_Div"
    df['Stoch_Div'] = div
    return df


def calculate_wma(series, window=21):
    weights = np.arange(1, window + 1, dtype=float)
    denom   = weights.sum()
    arr     = series.to_numpy(dtype=float)
    out     = np.full(len(arr), np.nan)
    if len(arr) >= window:
        windows    = sliding_window_view(arr, window)
        out[window - 1:] = windows @ weights / denom
    return pd.Series(out, index=series.index)


def calculate_lsma(series, window=25):
    n   = window
    arr = series.to_numpy(dtype=float)
    out = np.full(len(arr), np.nan)
    if len(arr) >= n:
        x      = np.arange(n, dtype=float)
        sum_x  = x.sum()
        sum_x2 = (x ** 2).sum()
        denom  = n * sum_x2 - sum_x ** 2

        windows  = sliding_window_view(arr, n)
        sum_y    = windows.sum(axis=1)
        sum_xy   = windows @ x

        slope     = (n * sum_xy - sum_x * sum_y) / denom
        intercept = (sum_y - slope * sum_x) / n
        out[n - 1:] = slope * (n - 1) + intercept
    return pd.Series(out, index=series.index)


def calculate_composite_signal(df, buy_threshold=2, sell_threshold=-2):
    # Score is kept purely for REFERENCE / sorting / cell colouring —
    # it does not decide Final_Signal (see rule-based logic below).
    score = pd.Series(0, index=df.index, dtype=int)

    score += np.select([df['Signal'] == 'LONG', df['Signal'] == 'SHORT'], [1, -1], default=0)
    score += np.select([df['Volume_Signal'] == 'Strong_Buy_Vol', df['Volume_Signal'] == 'Strong_Sell_Vol'], [1, -1], default=0)
    score += np.select([df['Stoch_Div'] == 'Bullish_Div', df['Stoch_Div'] == 'Bearish_Div'], [1, -1], default=0)
    score += np.select([df['BB_Position'] == 'Lower_Breakout', df['BB_Position'] == 'Upper_Breakout'], [1, -1], default=0)
    score += np.select([df['RSI'] < 30, df['RSI'] > 70], [1, -1], default=0)
    score += np.select([df['Diff_Trough'] == 'LONG', df['Diff_Peak'] == 'Short'], [1, -1], default=0)
    score += np.select([df['Gann_Reversal_Zone'] == 'Support', df['Gann_Reversal_Zone'] == 'Resistance'], [1, -1], default=0)

    df['Score'] = score

    # ── Rule-based Final_Signal ──────────────────────────────────────────────
    buy_trend    = (df['Diff_Trough'] == 'LONG') | (df['Signal'] == 'LONG')
    buy_confirm  = (df['Close_Color'] == 'Below_Support') | \
                   (df['BB_Position'] == 'Lower_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Buy_Vol')

    sell_trend   = (df['Diff_Peak'] == 'Short') | (df['Signal'] == 'SHORT')
    sell_confirm = (df['Close_Color'] == 'Above_Resistance') | \
                   (df['BB_Position'] == 'Upper_Breakout') | \
                   (df['Volume_Signal'] == 'Strong_Sell_Vol')

    df['Final_Signal'] = np.select(
        [buy_trend & buy_confirm, sell_trend & sell_confirm],
        ['BUY', 'SELL'],
        default='NEUTRAL'
    )
    return df


def calculate_signal(lsma, wma):
    above      = lsma > wma
    prev_above = above.shift(1).fillna(False)
    signal     = pd.Series("", index=lsma.index, dtype=str)
    signal[(~prev_above) & above]  = "LONG"
    signal[prev_above   & ~above]  = "SHORT"
    return signal


def add_close_color(df: pd.DataFrame) -> pd.DataFrame:
    """Above_Resistance (High > Gann_Resistance) / Below_Support (Low < Gann_Support) / Neutral."""
    df["Close_Color"] = np.where(
        df["HIGH"] > df["Gann_Resistance"],
        "Above_Resistance",
        np.where(
            df["LOW"] < df["Gann_Support"],
            "Below_Support",
            "Neutral"
        )
    )
    return df


# ── Excel export (identical to nifty50_analyzer.py, minus watchlist highlighting) ─
_GREEN_FILL = _RED_FILL = _GREEN_FONT = _RED_FONT = _HDR_FILL = _HDR_FONT = None
if OPENPYXL_OK:
    _GREEN_FILL = PatternFill(start_color="1A4731", end_color="1A4731", fill_type="solid")
    _RED_FILL   = PatternFill(start_color="4B1217", end_color="4B1217", fill_type="solid")
    _GREEN_FONT = Font(color="3FB950", bold=True)
    _RED_FONT   = Font(color="F85149", bold=True)
    _HDR_FILL   = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
    _HDR_FONT   = Font(color="8B949E", bold=True)


def save_excel_with_colors(df: pd.DataFrame, filepath: str) -> None:
    if not OPENPYXL_OK:
        print("  ⚠  openpyxl not available — skipping Excel export.")
        return

    from openpyxl.worksheet.filters import FilterColumn, Filters, DateGroupItem

    _signal_order = {"BUY": 0, "SELL": 1, "NEUTRAL": 2}
    df = df.copy()
    df["_sort_key"] = df["Final_Signal"].map(_signal_order).fillna(3)
    df["Date"]      = pd.to_datetime(df["Date"])
    df = df.sort_values(["_sort_key", "Date"], ascending=[True, False]).drop("_sort_key", axis=1).reset_index(drop=True)

    latest_date = df["Date"].max()

    wb = Workbook()
    ws = wb.active
    ws.title = "FnO Signals"

    cols            = list(df.columns)
    close_col_idx   = cols.index("CLOSE")       + 1
    close_color_idx = cols.index("Close_Color") + 1
    date_col_0      = cols.index("Date")

    for c_idx, col_name in enumerate(cols, 1):
        cell           = ws.cell(row=1, column=c_idx, value=col_name)
        cell.fill      = _HDR_FILL
        cell.font      = _HDR_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        color_label = row[close_color_idx - 1]
        for c_idx, val in enumerate(row, 1):
            col_name = cols[c_idx - 1]
            if col_name == "Date":
                dt = val.to_pydatetime() if hasattr(val, "to_pydatetime") else val
                write_val = dt.replace(tzinfo=None)
            elif hasattr(val, "isoformat"):
                write_val = str(val)[:19]
            else:
                write_val = val
            cell = ws.cell(row=r_idx, column=c_idx, value=write_val)
            if col_name == "Date":
                cell.number_format = "YYYY-MM-DD HH:MM:SS"
            if c_idx == close_col_idx:
                if color_label == "Above_Resistance":
                    cell.fill, cell.font = _RED_FILL, _RED_FONT
                elif color_label == "Below_Support":
                    cell.fill, cell.font = _GREEN_FILL, _GREEN_FONT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    fc = FilterColumn(colId=date_col_0)
    fc.filters = Filters(dateGroupItem=[DateGroupItem(
        year=int(latest_date.year), month=int(latest_date.month),
        day=int(latest_date.day), dateTimeGrouping="day",
    )])
    ws.auto_filter.filterColumn.append(fc)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 28)
    ws.freeze_panes = "A2"
    wb.save(filepath)


# ── Main Function — Kite-based fetch, same output schema as fetch_nifty50_data ─
def fetch_fno_data(kite, n_days: int = 30, interval: str = "1d",
                    buy_threshold: int = 2, sell_threshold: int = -2,
                    ticker_list=None, instruments_df: pd.DataFrame = None) -> pd.DataFrame:
    """
    kite           : an authenticated kiteconnect.KiteConnect instance.
    instruments_df : pass a pre-fetched/cached instrument master (from
                     get_instrument_master()) to avoid re-fetching it on
                     every call — recommended to cache this in the dashboard
                     with a ~1hr TTL, since it's a heavy call.
    """
    tickers = ticker_list if ticker_list is not None else FnO

    if interval not in KITE_INTERVAL_MAP:
        raise ValueError(f"Unsupported interval '{interval}'. Must be one of {list(KITE_INTERVAL_MAP)}.")
    kite_interval, resample_rule = KITE_INTERVAL_MAP[interval]

    if instruments_df is None:
        instruments_df = get_instrument_master(kite)
    if instruments_df.empty:
        raise ValueError("Could not load the Kite instrument master (empty). Check your access token / connection.")

    to_date   = now_ist()
    from_date = to_date - timedelta(days=n_days)

    all_frames = []
    for name in tickers:
        token, tsym, exch = resolve_instrument(instruments_df, name)
        if token is None:
            print(f"  ⚠ Could not resolve instrument for '{name}' — skipping "
                  f"(check INDEX_ALIASES/FUTURES_UNDERLYINGS if this is an index/commodity/currency).")
            continue

        try:
            candles = _historical_with_retry(kite, token, from_date, to_date, kite_interval)
        except Exception as e:
            print(f"  ✗ Error fetching {name} ({tsym}): {e}")
            continue
        time.sleep(0.34)  # stay under Kite's ~3 req/sec historical-data rate limit

        if not candles:
            print(f"  ⚠ No candles returned for {name} ({tsym})")
            continue

        try:
            df = pd.DataFrame(candles).rename(columns={
                "date": "Date", "open": "OPEN", "high": "HIGH",
                "low": "LOW", "close": "CLOSE", "volume": "VOLUME",
            })
            if resample_rule:
                df = _resample_candles(df, resample_rule)
            if df.empty:
                continue

            df["Stock Name"] = name
            df["Interval"]   = interval
            df["Return"]     = df["CLOSE"].diff().round(2)

            df["WMA"]      = calculate_wma(df["CLOSE"])
            df["LSMA"]     = calculate_lsma(df["CLOSE"])
            df["LSMA-WMA"] = (df["LSMA"] - df["WMA"]).round(2)
            df["Signal"]   = calculate_signal(df["LSMA"], df["WMA"])

            df["RSI"]              = calculate_rsi(df["CLOSE"])
            df["Stoch_K"], df["Stoch_D"] = calculate_stochastic(df)

            df["BB_Middle"], df["BB_Upper"], df["BB_Lower"], df["BB_Width"] = \
                calculate_bollinger_bands(df["CLOSE"])
            df["BB_Position"] = bollinger_position(df["CLOSE"], df["BB_Upper"], df["BB_Lower"])

            df = volume_analysis(df)
            df = simple_gann_time_signals(df)
            df = calculate_gann_targets(df)
            df = detect_stochastic_divergence(df)
            df = add_close_color(df)

            resistance_change = df["Gann_Resistance"].diff()
            support_change     = df["Gann_Support"].diff()
            level_shift = pd.Series("", index=df.index, dtype=str)
            level_shift[(resistance_change < 0) | (support_change < 0)] = "Level_Dropped"
            level_shift[support_change > 0] = "Support_Up"
            df["Gann_Level_Shift"] = level_shift

            band_width     = df["Gann_Resistance"] - df["Gann_Support"]
            resistance_pct = (df["Gann_Resistance"] - df["CLOSE"]) / band_width * 100
            support_pct    = (df["CLOSE"] - df["Gann_Support"]) / band_width * 100
            gap_vals = []
            for cc, rp, sp in zip(df["Close_Color"], resistance_pct, support_pct):
                if cc == "Neutral" and pd.notna(rp) and pd.notna(sp):
                    gap_vals.append(f"{rp:.2f}% // {sp:.2f}%")
                else:
                    gap_vals.append("")
            df["Resistance_Support_Gap"] = gap_vals

            diff = df["LSMA-WMA"]
            df["Diff_Peak"]   = ((diff < diff.shift(1)) & (diff.shift(1) > diff.shift(2))).map({True: "Short",  False: ""})
            df["Diff_Trough"] = ((diff > diff.shift(1)) & (diff.shift(1) < diff.shift(2))).map({True: "LONG", False: ""})

            df = calculate_composite_signal(df, buy_threshold=buy_threshold, sell_threshold=sell_threshold)
            all_frames.append(df)

        except Exception as e:
            print(f"  ✗ Error processing {name} ({tsym}): {e}")

    if not all_frames:
        raise ValueError("No data fetched for any instrument. Check your token, instrument resolution, and connection.")

    result = pd.concat(all_frames, ignore_index=True)

    cols = [
        "Stock Name", "Date", "Interval", "Score", "Final_Signal", "Resistance_Support_Gap", "Return",
        "OPEN", "HIGH", "LOW", "CLOSE", "Close_Color", "VOLUME",
        "RSI", "Stoch_K", "Stoch_D", "Stoch_Div",
        "BB_Middle", "BB_Upper", "BB_Lower", "BB_Width", "BB_Position",
        "Volume_SMA", "Volume_Ratio", "OBV", "Volume_Trend", "Volume_Signal",
        "WMA", "LSMA", "LSMA-WMA", "Signal",
        "Gann_Time", "Gann_Resistance", "Gann_Support", "Gann_Reversal_Zone", "Gann_Level_Shift",
        "Diff_Peak", "Diff_Trough",
    ]

    numeric_cols = result.select_dtypes(include=[np.number]).columns
    result[numeric_cols] = result[numeric_cols].round(2)
    result = result[cols]
    result = result.dropna(subset=["WMA", "LSMA", "RSI", "BB_Middle"]).reset_index(drop=True)
    return result
